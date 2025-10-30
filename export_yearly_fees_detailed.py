"""
Export Detailed Yearly Fees Calculations to Excel

Creates a comprehensive Excel workbook with:
1. Methodology Explanation sheet - Clear documentation of the calculation approach
2. Scenario Configuration sheet - Fee structure parameters
3. Yearly Summary sheet - Annual totals matching the PDF report
4. Monthly Detail sheet - All monthly calculations showing every step
5. Daily Detail sheet - All daily calculations showing every step
6. Market Returns (Daily) sheet - Individual market returns showing how daily totals are calculated

This allows the manager to review and verify the calculation methodology.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from database import Database
from fee_scenarios import load_fee_scenario, calculate_net_nav_series, get_daily_returns
from datetime import date
from typing import Optional


def create_explanation_sheet(wb, scenario):
    """Create a sheet explaining the fee calculation methodology."""
    ws = wb.create_sheet("Methodology Explanation", 0)

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80

    # Title
    ws['A1'] = "Fee Calculation Methodology"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:B1')

    row = 3

    # Overview
    ws[f'A{row}'] = "Overview"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    ws[f'B{row}'] = (
        "This workbook implements a percentage-based fee calculation methodology. "
        "All fees are calculated as percentages of returns (not dollar amounts on NAV). "
        "Returns are additive (summed), not compounded."
    )
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 2

    # Key Concepts
    ws[f'A{row}'] = "Key Concepts"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    concepts = [
        ("Deflation Factor", f"{scenario.deflation_factor} - All raw returns are multiplied by this factor before fee calculations. "
                           f"This represents a {(1-scenario.deflation_factor)*100:.1f}% reduction in returns."),
        ("Cumulative Returns", "Returns are summed additively, NOT compounded geometrically. "
                              "Cumulative Return = Sum of all daily deflated returns."),
        ("High Water Mark", "Tracks the highest cumulative return percentage achieved. "
                          "Performance fees are only charged on profits ABOVE the previous high water mark."),
        ("Rolling 12-Month Return", f"Used for performance fee tier determination. "
                                   f"Calculated as the sum of the last 261 trading days of deflated returns."),
        ("Management Fee", f"{scenario.management_bands[0].annual_percentage*100:.2f}% annual = "
                         f"{scenario.management_bands[0].annual_percentage/12*100:.4f}% monthly. "
                         f"Applied every month regardless of performance."),
    ]

    if scenario.performance_bands and len(scenario.performance_bands) == 2:
        low = scenario.performance_bands[0]
        high = scenario.performance_bands[1]
        concepts.append((
            "Performance Fee (Interpolated)",
            f"Linear interpolation between {low.fee_percentage*100:.0f}% and {high.fee_percentage*100:.0f}%. "
            f"At {low.performance_min*100:.0f}% rolling 12-month return → {low.fee_percentage*100:.0f}% fee. "
            f"At {high.performance_min*100:.0f}% rolling 12-month return → {high.fee_percentage*100:.0f}% fee. "
            f"Between these points, the fee rate increases linearly."
        ))

    for concept, description in concepts:
        ws[f'A{row}'] = concept
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # Calculation Steps
    ws[f'A{row}'] = "Daily Calculation Steps"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    daily_steps = [
        ("1. Raw Return", "Daily return from trading (sum across all non-benchmark markets)"),
        ("2. Apply Deflation", f"Deflated Return = Raw Return × {scenario.deflation_factor}"),
        ("3. Update Cumulative", "Cumulative Return = Previous Cumulative + Deflated Return (additive)"),
        ("4. Update HWM", "High Water Mark = MAX(High Water Mark, Cumulative Return)"),
    ]

    for step, description in daily_steps:
        ws[f'A{row}'] = step
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        row += 1

    row += 1

    # Monthly Calculation Steps
    ws[f'A{row}'] = "Monthly Calculation Steps"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    monthly_steps = [
        ("1. Monthly Return", "Change in cumulative return this month = Current Cumulative - Previous Cumulative"),
        ("2. Profit Above HWM", "MAX(0, Current Cumulative - Previous HWM)"),
        ("3. Rolling 12-Month Return", "Sum of last 261 trading days of deflated returns"),
        ("4. Collar Performance", f"Clamp rolling return between {scenario.performance_bands[0].performance_min*100:.0f}% and {scenario.performance_bands[-1].performance_min*100:.0f}%"),
        ("5. Performance Fee Rate", "Calculate using linear interpolation based on collared performance"),
        ("6. Management Fee", f"{scenario.management_bands[0].annual_percentage/12*100:.4f}% (applied monthly)"),
        ("7. Performance Fee", "Performance Fee Rate × Profit Above HWM"),
        ("8. Total Fee", "Management Fee + Performance Fee"),
    ]

    for step, description in monthly_steps:
        ws[f'A{row}'] = step
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = description
        row += 1

    row += 1

    # Example Calculation
    ws[f'A{row}'] = "Example Calculation"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    ws[f'B{row}'] = (
        "EXAMPLE: If rolling 12-month return is 40%:\n"
        f"Performance Fee Rate = {scenario.performance_bands[0].fee_percentage*100:.0f}% + "
        f"({scenario.performance_bands[-1].fee_percentage*100:.0f}% - {scenario.performance_bands[0].fee_percentage*100:.0f}%) × "
        f"(40% - {scenario.performance_bands[0].performance_min*100:.0f}%) / "
        f"({scenario.performance_bands[-1].performance_min*100:.0f}% - {scenario.performance_bands[0].performance_min*100:.0f}%)\n"
        f"= {scenario.performance_bands[0].fee_percentage*100:.0f}% + "
        f"{(scenario.performance_bands[-1].fee_percentage - scenario.performance_bands[0].fee_percentage)*100:.0f}% × "
        f"(20%) / (40%)\n"
        f"= {scenario.performance_bands[0].fee_percentage*100:.0f}% + {(scenario.performance_bands[-1].fee_percentage - scenario.performance_bands[0].fee_percentage)*100:.0f}% × 0.5\n"
        f"= {scenario.performance_bands[0].fee_percentage*100:.0f}% + {(scenario.performance_bands[-1].fee_percentage - scenario.performance_bands[0].fee_percentage)*100/2:.0f}%\n"
        f"= 30%"
    )
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 90

    # Freeze panes
    ws.freeze_panes = 'A2'


def create_scenario_config_sheet(wb, scenario, fund_size):
    """Create a sheet showing the scenario configuration."""
    ws = wb.create_sheet("Scenario Configuration")

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50

    # Title
    ws['A1'] = "Fee Scenario Configuration"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:B1')

    row = 3

    # Basic Info
    config_data = [
        ("Scenario Name", scenario.name),
        ("Description", scenario.description),
        ("Fund Size (USD)", f"${fund_size:,.0f}"),
        ("", ""),
        ("Deflation Factor", scenario.deflation_factor),
        ("Deflation Percentage", f"{(1-scenario.deflation_factor)*100:.1f}%"),
        ("", ""),
        ("Management Fee (Annual)", f"{scenario.management_bands[0].annual_percentage*100:.2f}%"),
        ("Management Fee (Monthly)", f"{scenario.management_bands[0].annual_percentage/12*100:.4f}%"),
        ("", ""),
        ("Performance Tier Basis", scenario.performance_tier_basis.replace('_', ' ').title()),
        ("High Water Mark", "Yes" if scenario.use_high_water_mark else "No"),
    ]

    for label, value in config_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    row += 1
    ws[f'A{row}'] = "Performance Fee Tiers"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    # Performance fee tiers header
    ws[f'A{row}'] = "Return Threshold"
    ws[f'B{row}'] = "Fee Percentage"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 1

    for band in scenario.performance_bands:
        ws[f'A{row}'] = f"{band.performance_min*100:.0f}%"
        ws[f'B{row}'] = f"{band.fee_percentage*100:.0f}%"
        row += 1

    row += 1
    ws[f'A{row}'] = "Interpolation Type"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = scenario.performance_bands[0].interpolation_type.title()


def create_monthly_detail_sheet(wb, monthly_series, scenario):
    """Create detailed monthly calculations sheet."""
    ws = wb.create_sheet("Monthly Detail")

    # Headers
    headers = [
        "Month", "Monthly\nReturn (%)", "Cumulative\nReturn (%)", "High Water\nMark (%)",
        "Profit Above\nHWM (%)", "Rolling 12mo\nReturn (%)", "Collared\nPerformance (%)",
        "Mgmt Fee (%)", "Perf Fee\nRate (%)", "Perf Fee (%)", "Total Fee (%)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths
    widths = [12, 12, 12, 12, 12, 12, 12, 10, 10, 10, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data rows
    for row_idx, calc in enumerate(monthly_series, 2):
        ws.cell(row=row_idx, column=1, value=calc.date.strftime('%Y-%m'))
        ws.cell(row=row_idx, column=2, value=calc.monthly_return_pct * 100)
        ws.cell(row=row_idx, column=3, value=calc.cumulative_return_pct * 100)
        ws.cell(row=row_idx, column=4, value=calc.high_water_mark_pct * 100)
        ws.cell(row=row_idx, column=5, value=calc.profit_above_hwm_pct * 100)
        ws.cell(row=row_idx, column=6, value=calc.rolling_12mo_return_pct * 100)
        ws.cell(row=row_idx, column=7, value=calc.collared_performance * 100)
        ws.cell(row=row_idx, column=8, value=calc.management_fee_pct * 100)
        ws.cell(row=row_idx, column=9, value=calc.performance_fee_rate * 100)
        ws.cell(row=row_idx, column=10, value=calc.performance_fee_pct * 100)
        ws.cell(row=row_idx, column=11, value=calc.total_fee_pct * 100)

        # Format numbers
        for col in range(2, 12):
            ws.cell(row=row_idx, column=col).number_format = '0.0000'

        # Alternating row colors
        if row_idx % 2 == 0:
            for col in range(1, 12):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
                )

    # Freeze panes
    ws.freeze_panes = 'A2'


def create_market_returns_sheet(wb, db, program_id, start_date, end_date):
    """Create sheet showing individual market returns for each day."""
    ws = wb.create_sheet("Market Returns (Daily)")

    # Get all non-benchmark markets for this program
    markets_query = """
        SELECT DISTINCT m.id, m.name
        FROM markets m
        JOIN pnl_records pr ON pr.market_id = m.id
        WHERE pr.program_id = ?
          AND pr.resolution = 'daily'
          AND m.is_benchmark = 0
        ORDER BY m.name
    """
    markets = db.fetch_all(markets_query, (program_id,))

    if not markets:
        print("Warning: No markets found")
        return

    market_ids = [m['id'] for m in markets]
    market_names = [m['name'] for m in markets]

    # Headers: Date + each market name + Total
    headers = ['Date'] + market_names + ['Total (%)']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths
    ws.column_dimensions['A'].width = 12  # Date column
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # Get all daily returns
    returns_query = """
        SELECT pr.date, pr.market_id, pr.return
        FROM pnl_records pr
        JOIN markets m ON pr.market_id = m.id
        WHERE pr.program_id = ?
          AND pr.resolution = 'daily'
          AND pr.date >= ?
          AND pr.date <= ?
          AND m.is_benchmark = 0
        ORDER BY pr.date, m.name
    """

    all_returns = db.fetch_all(returns_query, (program_id, start_date.isoformat(), end_date.isoformat()))

    # Organize by date
    from collections import defaultdict
    returns_by_date = defaultdict(dict)

    for record in all_returns:
        trade_date = record['date']
        market_id = record['market_id']
        return_val = record['return']
        returns_by_date[trade_date][market_id] = return_val

    # Write data rows
    row_idx = 2
    for trade_date in sorted(returns_by_date.keys()):
        # Date column
        ws.cell(row=row_idx, column=1, value=trade_date)

        # Market columns
        total_return = 0.0
        for col_idx, market_id in enumerate(market_ids, 2):
            return_pct = returns_by_date[trade_date].get(market_id, 0.0) * 100
            ws.cell(row=row_idx, column=col_idx, value=return_pct)
            ws.cell(row=row_idx, column=col_idx).number_format = '0.0000'
            total_return += returns_by_date[trade_date].get(market_id, 0.0)

        # Total column
        ws.cell(row=row_idx, column=len(headers), value=total_return * 100)
        ws.cell(row=row_idx, column=len(headers)).number_format = '0.0000'
        ws.cell(row=row_idx, column=len(headers)).font = Font(bold=True)

        # Alternating row colors
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
                )

        row_idx += 1

    # Freeze panes (freeze first row and first column)
    ws.freeze_panes = 'B2'

    print(f"Added market returns for {len(returns_by_date)} days across {len(market_names)} markets")


def create_daily_detail_sheet(wb, db, program_id, scenario, start_date, end_date):
    """Create detailed daily calculations sheet with all steps."""
    ws = wb.create_sheet("Daily Detail")

    # Headers
    headers = [
        "Date", "Raw\nReturn (%)", "Deflated\nReturn (%)", "Cumulative\nReturn (%)",
        "High Water\nMark (%)", "Rolling 261d\nSum (%)", "Days Since\nInception"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths
    widths = [12, 12, 12, 12, 12, 12, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Get daily returns
    daily_returns_df = get_daily_returns(db, program_id, start_date, end_date)

    if len(daily_returns_df) == 0:
        print("Warning: No daily returns found")
        return

    # Calculate all daily values
    cumulative = 0.0
    hwm = 0.0
    rolling_window = []

    for row_idx, row in enumerate(daily_returns_df.itertuples(index=False), 2):
        trade_date = row.date
        raw_return = row.daily_return
        # Apply deflation
        deflated_return = raw_return * scenario.deflation_factor

        # Update cumulative (additive)
        cumulative += deflated_return

        # Update high water mark
        if cumulative > hwm:
            hwm = cumulative

        # Update rolling window (last 261 days)
        rolling_window.append(deflated_return)
        if len(rolling_window) > 261:
            rolling_window.pop(0)

        rolling_sum = sum(rolling_window)
        days_since_inception = row_idx - 1

        # Write row
        ws.cell(row=row_idx, column=1, value=trade_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_idx, column=2, value=raw_return * 100)
        ws.cell(row=row_idx, column=3, value=deflated_return * 100)
        ws.cell(row=row_idx, column=4, value=cumulative * 100)
        ws.cell(row=row_idx, column=5, value=hwm * 100)
        ws.cell(row=row_idx, column=6, value=rolling_sum * 100)
        ws.cell(row=row_idx, column=7, value=days_since_inception)

        # Format numbers
        for col in range(2, 7):
            ws.cell(row=row_idx, column=col).number_format = '0.0000'

        # Alternating row colors
        if row_idx % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
                )

    # Freeze panes
    ws.freeze_panes = 'A2'

    print(f"Added {len(daily_returns_df)} daily records to Daily Detail sheet")


def create_yearly_summary_sheet(wb, monthly_series, fund_size, start_date, end_date):
    """Create yearly summary sheet matching the PDF report."""
    ws = wb.create_sheet("Yearly Summary")

    # Headers
    headers = ["Year", "Mgmt Fee ($)", "Perf Fee ($)", "Total Fee ($)",
               "Investor Profit ($)", "CAGR (%)"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 8
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['F'].width = 12

    # Group by year
    yearly_data = {}
    previous_cumulative = 0.0
    previous_hwm = 0.0

    for i, calc in enumerate(monthly_series):
        year = calc.date.year

        if year not in yearly_data:
            yearly_data[year] = {
                'start_cumulative': previous_cumulative,
                'start_hwm': previous_hwm,
                'months': []
            }

        yearly_data[year]['months'].append(calc)
        previous_cumulative = calc.cumulative_return_pct
        previous_hwm = calc.high_water_mark_pct

    # Calculate yearly aggregates
    inception_date = monthly_series[0].date
    row_idx = 2

    for year in sorted(yearly_data.keys()):
        year_info = yearly_data[year]
        months = year_info['months']

        # Calculate totals
        mgmt_fee_pct = sum(m.management_fee_pct for m in months)
        perf_fee_pct = sum(m.performance_fee_pct for m in months)
        total_fee_pct = sum(m.total_fee_pct for m in months)

        # Convert to USD
        mgmt_fee_usd = mgmt_fee_pct * fund_size
        perf_fee_usd = perf_fee_pct * fund_size
        total_fee_usd = total_fee_pct * fund_size

        # Calculate investor values
        end_cumulative = months[-1].cumulative_return_pct
        year_return_pct = end_cumulative - year_info['start_cumulative']
        investor_year_profit = (year_return_pct - total_fee_pct) * fund_size

        # Calculate cumulative CAGR from inception
        months_since_inception = monthly_series.index(months[-1]) + 1
        total_fees_to_date = sum(m.total_fee_pct for m in monthly_series[:months_since_inception])
        investor_net_return = end_cumulative - total_fees_to_date
        investor_value = fund_size * (1.0 + investor_net_return)

        years_from_inception = (months[-1].date - inception_date).days / 365.25
        if years_from_inception > 0 and investor_value > 0:
            cagr = ((investor_value / fund_size) ** (1.0 / years_from_inception) - 1.0) * 100
        else:
            cagr = 0.0

        # Write row
        ws.cell(row=row_idx, column=1, value=year)
        ws.cell(row=row_idx, column=2, value=mgmt_fee_usd)
        ws.cell(row=row_idx, column=3, value=perf_fee_usd)
        ws.cell(row=row_idx, column=4, value=total_fee_usd)
        ws.cell(row=row_idx, column=5, value=investor_year_profit)
        ws.cell(row=row_idx, column=6, value=cagr)

        # Format
        for col in range(2, 6):
            ws.cell(row=row_idx, column=col).number_format = '#,##0'
        ws.cell(row=row_idx, column=6).number_format = '0.0'

        # Alternating colors
        if row_idx % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
                )

        row_idx += 1

    # Add totals row
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="TOTAL")
    ws.cell(row=row_idx, column=1).font = Font(bold=True)

    total_mgmt = sum(m.management_fee_pct for m in monthly_series) * fund_size
    total_perf = sum(m.performance_fee_pct for m in monthly_series) * fund_size
    total_fees = sum(m.total_fee_pct for m in monthly_series) * fund_size

    ws.cell(row=row_idx, column=2, value=total_mgmt)
    ws.cell(row=row_idx, column=3, value=total_perf)
    ws.cell(row=row_idx, column=4, value=total_fees)

    for col in range(2, 5):
        ws.cell(row=row_idx, column=col).number_format = '#,##0'
        ws.cell(row=row_idx, column=col).font = Font(bold=True)
        ws.cell(row=row_idx, column=col).fill = PatternFill(
            start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
        )

    # Freeze panes
    ws.freeze_panes = 'A2'


def export_yearly_fees_to_excel(
    db,
    program_id: int,
    scenario_id: int,
    output_path: str,
    fund_size: float = 10_000_000,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
):
    """
    Export detailed yearly fees calculations to Excel.

    Creates a workbook with multiple sheets:
    - Methodology Explanation
    - Scenario Configuration
    - Yearly Summary (matching PDF)
    - Monthly Detail (all monthly calculation steps)
    - Daily Detail (all daily calculation steps)
    - Market Returns (Daily) (individual market returns for verification)
    """
    # Get program info
    program = db.fetch_one("""
        SELECT p.program_name, p.starting_date, m.manager_name
        FROM programs p
        JOIN managers m ON p.manager_id = m.id
        WHERE p.id = ?
    """, (program_id,))

    if not program:
        raise ValueError(f"Program {program_id} not found")

    # Load scenario
    scenario = load_fee_scenario(db, scenario_id)

    # Determine date range
    if start_date is None:
        start_date = date.fromisoformat(program['starting_date'])

    if end_date is None:
        latest = db.fetch_one("""
            SELECT MAX(date) as max_date
            FROM pnl_records
            WHERE program_id = ? AND resolution = 'daily'
        """, (program_id,))
        end_date = date.fromisoformat(latest['max_date']) if latest and latest['max_date'] else date.today()

    # Calculate monthly series
    print(f"Calculating fees for {program['manager_name']} {program['program_name']}...")
    monthly_series = calculate_net_nav_series(
        db, program_id, scenario, start_date, end_date, fund_size
    )

    if len(monthly_series) == 0:
        raise ValueError("No data returned from calculations")

    print(f"Processed {len(monthly_series)} months")

    # Create workbook
    print(f"Creating Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create all sheets
    create_explanation_sheet(wb, scenario)
    create_scenario_config_sheet(wb, scenario, fund_size)
    create_yearly_summary_sheet(wb, monthly_series, fund_size, start_date, end_date)
    create_monthly_detail_sheet(wb, monthly_series, scenario)
    create_daily_detail_sheet(wb, db, program_id, scenario, start_date, end_date)
    create_market_returns_sheet(wb, db, program_id, start_date, end_date)

    # Save
    wb.save(output_path)
    print(f"Saved: {output_path}")

    return output_path


def main():
    """Export detailed yearly fees for Alpha Bet MFT."""
    with Database() as db:
        # Get Alpha Bet MFT program
        program = db.fetch_one("""
            SELECT p.id
            FROM programs p
            JOIN managers m ON p.manager_id = m.id
            WHERE m.manager_name = 'Alpha Bet' AND p.program_name = 'MFT'
        """)

        if not program:
            print("Error: Alpha Bet MFT program not found")
            return

        # Get Alpha Bet Standard scenario
        scenario = db.fetch_one("""
            SELECT id FROM fee_scenarios WHERE scenario_name = 'Alpha Bet Standard'
        """)

        if not scenario:
            print("Error: Alpha Bet Standard scenario not found")
            return

        # Export to Excel
        export_yearly_fees_to_excel(
            db,
            program_id=program['id'],
            scenario_id=scenario['id'],
            output_path='export/alpha_bet/mft/fee_scenarios/alpha_bet_mft_yearly_fees_detailed.xlsx',
            fund_size=10_000_000,
            start_date=date(2006, 1, 3),
            end_date=date(2025, 10, 17)
        )


if __name__ == "__main__":
    main()
